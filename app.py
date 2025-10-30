
import streamlit as st
import pandas as pd
import numpy as np
import math
from io import BytesIO
import folium
from streamlit_folium import st_folium

# Elevación offline-friendly con SRTM (tiles HGT cacheados automáticamente)
# pip install srtm.py
import srtm

st.set_page_config(page_title="Δh – Rugosidad (Norma FM Panamá) – SRTM con paso configurable", layout="wide")

# ---------------- Geo helpers ----------------
R_EARTH_M = 6371000.0

def destination_point(lat_deg, lon_deg, bearing_deg, distance_m):
    lat1 = math.radians(lat_deg)
    lon1 = math.radians(lon_deg)
    brng = math.radians(bearing_deg)
    dr = distance_m / R_EARTH_M
    lat2 = math.asin(math.sin(lat1) * math.cos(dr) + math.cos(lat1) * math.sin(dr) * math.cos(brng))
    lon2 = lon1 + math.atan2(math.sin(brng) * math.sin(dr) * math.cos(lat1),
                             math.cos(dr) - math.sin(lat1) * math.sin(lat2))
    return math.degrees(lat2), (math.degrees(lon2) + 540) % 360 - 180

def compute_delta_h(elevations_m):
    arr = np.array([e for e in elevations_m if e is not None], dtype=float)
    if arr.size == 0:
        return None, None, None
    h10 = float(np.percentile(arr, 10))
    h90 = float(np.percentile(arr, 90))
    delta_h = h10 - h90
    return delta_h, h10, h90

def deltaF_from_deltaH(delta_h, freq_mhz):
    # Norma Panamá: ΔF = 1.9 - 0.03 * Δh * (1 + f/300)
    return 1.9 - 0.03 * (delta_h) * (1.0 + freq_mhz/300.0)

# ---------------- UI ----------------
st.title("🗺️ Índice de rugosidad Δh (Norma FM – Panamá) – Elevación SRTM (paso configurable)")
st.caption("Tramo normativo: 10–50 km desde la antena. Puedes elegir el intervalo de muestreo (norma: 500 m).")

with st.sidebar:
    st.header("Parámetros")
    lat = st.number_input("Latitud de la antena (°)", value=8.806600, format="%.6f")
    lon = st.number_input("Longitud de la antena (°)", value=-82.540300, format="%.6f")
    freq = st.number_input("Frecuencia (MHz) para ΔF", value=100.0, min_value=50.0, max_value=300.0, step=0.1, format="%.1f")
    st.markdown("---")
    mode = st.radio("Modo de azimuts", ["Un solo azimut", "Varios azimuts"], index=1)
    if mode == "Un solo azimut":
        az_list = [st.number_input("Azimut (°)", value=0.0, min_value=0.0, max_value=359.9, step=0.1)]
    else:
        az_str = st.text_input("Lista de azimuts (°) separados por coma", value="0,45,90,135,180,225,270,315")
        try:
            az_list = [float(x.strip()) for x in az_str.split(",") if x.strip()!=""]
        except Exception:
            az_list = []
            st.error("Formato de lista inválido. Usa valores numéricos separados por coma.")
    st.markdown("---")
    start_km = 10.0
    end_km = 50.0
    step_m = st.number_input(
        "Distancia entre puntos (m)",
        min_value=100, max_value=2000, value=500, step=50,
        help="Intervalo de muestreo. La norma utiliza 500 m."
    )
    if step_m == 500:
        st.success("✅ Cumple con el muestreo de la norma (500 m).")
    elif step_m < 500:
        st.warning("⚠️ Paso más fino que la norma (ok para análisis). Para informes normativos usa 500 m.")
    else:
        st.warning("⚠️ Paso más grueso que la norma. Para informes normativos usa 500 m.")
    st.markdown("---")
    st.checkbox("Guardar resultados en sesión (persistir)", value=True, key="persist_results")

if "runs" not in st.session_state:
    st.session_state.runs = []  # lista de DataFrames de resultados por ejecución

col1, col2 = st.columns([1,1])
with col1:
    run = st.button("Calcular Δh")
with col2:
    clear = st.button("Limpiar historial")

if clear:
    st.session_state.runs = []
    st.success("Historial borrado.")

# Preparar SRTM (carga/caché)
@st.cache_resource
def get_srtm_data():
    return srtm.get_data()

def get_elevations_srtm(lat_list, lon_list, data):
    elevs = []
    for la, lo in zip(lat_list, lon_list):
        elevs.append(data.get_elevation(la, lo))
    return elevs

if run and len(az_list) == 0:
    st.error("Debes especificar al menos un azimut.")
    st.stop()

if run:
    data = get_srtm_data()
    results = []
    map_center = (lat, lon)
    fmap = folium.Map(location=map_center, zoom_start=8, control_scale=True)
    profiles_store = {}  # perfiles detallados por azimut

    # Construir distancias según paso
    distances_m = list(range(int(start_km*1000), int(end_km*1000)+1, int(step_m)))
    n_points = len(distances_m)

    for az in az_list:
        lats, lons = [], []
        for d in distances_m:
            plat, plon = destination_point(lat, lon, az, d)
            lats.append(plat); lons.append(plon)

        elev = get_elevations_srtm(lats, lons, data)

        delta_h, h10, h90 = compute_delta_h(elev)
        if delta_h is None:
            st.warning(f"Sin datos de elevación para azimut {az}°")
            continue

        dF = deltaF_from_deltaH(delta_h, freq)
        row = {
            "Azimut (°)": az,
            "Δh (m)": round(delta_h, 2),
            "h10 (m)": round(h10, 2),
            "h90 (m)": round(h90, 2),
            "ΔF (dB)": round(dF, 2),
            "Puntos muestreados": n_points,
            "Paso (m)": int(step_m),
        }
        results.append(row)

        profiles_store[az] = pd.DataFrame({
            "Distancia (km)": [d/1000.0 for d in distances_m],
            "Lat": lats,
            "Lon": lons,
            "Elevación (m)": elev
        })

        folium.PolyLine(list(zip(lats, lons)), weight=3, opacity=0.7).add_to(fmap)

    if len(results)==0:
        st.stop()

    res_df = pd.DataFrame(results).sort_values("Azimut (°)").reset_index(drop=True)

    if st.session_state.persist_results:
        st.session_state.runs.append(res_df.copy())

    st.subheader("Resultados por azimut (ejecución actual)")
    st.dataframe(res_df, use_container_width=True)

    st.markdown("**Resumen (promedios de esta ejecución):**")
    st.write({
        "Δh promedio (m)": round(res_df["Δh (m)"].mean(), 2),
        "ΔF promedio (dB)": round(res_df["ΔF (dB)"].mean(), 2),
        "Puntos promedio": int(res_df["Puntos muestreados"].mean())
    })

    folium.Marker(location=map_center, tooltip="Transmisor").add_to(fmap)
    st_folium(fmap, width=None, height=520)

    # ---- Descargas de la ejecución ----
    def to_excel_bytes(df):
        from openpyxl import Workbook
        from openpyxl.utils.dataframe import dataframe_to_rows
        wb = Workbook()
        ws = wb.active
        ws.title = "DeltaH"
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        ws["I1"] = "Δh Norma FM Panamá: Δh = h10 - h90 (10–50 km)"
        ws["I2"] = "Muestreo cada 500 m (norma). Este archivo puede usar otro paso si el usuario lo modificó."
        ws["I3"] = "ΔF = 1.9 - 0.03*Δh*(1 + f/300)"
        out = BytesIO()
        wb.save(out)
        return out.getvalue()

    csv_bytes = res_df.to_csv(index=False).encode("utf-8")
    xlsx_bytes = to_excel_bytes(res_df)

    st.download_button("⬇️ Descargar CSV (ejecución)", data=csv_bytes, file_name="deltaH_resultados.csv", mime="text/csv")
    st.download_button("⬇️ Descargar Excel (ejecución)", data=xlsx_bytes, file_name="deltaH_resultados.xlsx",
                       mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    # Perfiles por azimut (ZIP de CSV)
    with BytesIO() as zip_buffer:
        import zipfile
        with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zf:
            for az, prof in profiles_store.items():
                zf.writestr(f"perfil_azimut_{az:.1f}.csv", prof.to_csv(index=False))
        st.download_button("⬇️ Descargar perfiles (ZIP)", data=zip_buffer.getvalue(),
                           file_name="perfiles_radiales.zip", mime="application/zip")

# ---- Historial de resultados (sesión) ----
if len(st.session_state.runs) > 0:
    st.subheader("Historial de resultados (sesión)")
    frames = []
    for idx, df in enumerate(st.session_state.runs, start=1):
        tmp = df.copy()
        tmp.insert(0, "Ejecución #", idx)
        frames.append(tmp)
    hist = pd.concat(frames, ignore_index=True)
    st.dataframe(hist, use_container_width=True)

    hist_csv = hist.to_csv(index=False).encode("utf-8")
    st.download_button("⬇️ Descargar CSV (historial)", data=hist_csv, file_name="deltaH_historial.csv", mime="text/csv")

st.caption("Norma FM Panamá: Δh = h10 - h90 calculado entre 10–50 km; muestreo recomendado: 500 m; ΔF = 1.9 - 0.03Δh(1+f/300).")
